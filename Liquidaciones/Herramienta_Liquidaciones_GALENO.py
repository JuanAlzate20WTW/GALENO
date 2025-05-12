import os
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import locale

# Configurar el idioma a español
locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')  # Para sistemas basados en Unix/Linux
# Para Windows, usa: locale.setlocale(locale.LC_TIME, 'Spanish_Spain.1252')

def limpiar_montos(df, columnas):
    for col in columnas:
        if col in df.columns:  # Verificar si la columna existe
            df[col] = df[col].astype(str) \
                             .str.replace('$', '', regex=False) \
                             .str.replace('.', '', regex=False) \
                             .str.replace(',', '.', regex=False) \
                             .str.strip()
            df[col] = pd.to_numeric(df[col], errors='coerce')
        else:
            print(f"Advertencia: La columna '{col}' no existe en el DataFrame.")
    return df

def ajustar_ancho_columna(filepath, columna, ancho):
    wb = load_workbook(filepath)
    ws = wb.active
    col_letter = get_column_letter(columna)
    ws.column_dimensions[col_letter].width = ancho
    wb.save(filepath)

def procesar_archivos_excel(carpeta_entrada, carpeta_salida):
    # Crear carpeta de salida si no existe
    os.makedirs(carpeta_salida, exist_ok=True)

    # Obtener mes y año actuales
    fecha_actual = datetime.now()
    fecha_anterior = fecha_actual - timedelta(days=fecha_actual.day)  # Restar días para obtener el mes anterior
    mes_anterior = fecha_anterior.strftime('%B')  # Nombre del mes anterior en español
    anio_anterior = fecha_anterior.year

    # Lista para almacenar los datos combinados
    datos_combinados = []

    # Procesar todos los archivos .xls en la carpeta de entrada
    for i, archivo in enumerate(os.listdir(carpeta_entrada)):
        if archivo.lower().endswith('.xls'):
            ruta_entrada = os.path.join(carpeta_entrada, archivo)
            print(f"Procesando: {ruta_entrada}")

            # Leer el archivo .xls
            try:
                df = pd.read_excel(ruta_entrada, engine='xlrd', header=1)  # Cambia `header=1` si los encabezados están en otra fila
                print(f"Nombres de columnas en {archivo}: {df.columns.tolist()}")

                # Normalizar nombres de columnas
                df.columns = df.columns.str.strip().str.lower().str.replace('á', 'a').str.replace('ó', 'o')

                # Seleccionar solo las columnas relevantes
                columnas_relevantes = ['contrato/poliza', 'cuit/cuil', 'importe cobranzas', 'comision legajo']
                df = df[columnas_relevantes]

                # Agregar los datos al conjunto combinado
                datos_combinados.append(df)
            except Exception as e:
                print(f"Error al leer {archivo}: {e}")
                continue

    # Combinar todos los datos en un único DataFrame
    if datos_combinados:
        df_combinado = pd.concat(datos_combinados, ignore_index=True)

        # Eliminar la primera fila si está vacía
        if df_combinado.iloc[0].isnull().all():
            df_combinado = df_combinado.iloc[1:]

        # Eliminar la última fila si contiene totales irrelevantes
        if df_combinado.iloc[-1].isnull().any() or df_combinado.iloc[-1].str.contains('total', case=False, na=False).any():
            df_combinado = df_combinado.iloc[:-1]

        # Limpiar columnas numéricas
        columnas_a_limpiar = ['importe cobranzas', 'comision legajo']
        df_combinado = limpiar_montos(df_combinado, columnas_a_limpiar)

        # Generar nombre de salida .xlsx con el mes anterior
        nombre_salida = f"GALENO Comision Final - {mes_anterior.capitalize()} {anio_anterior}.xlsx"
        ruta_salida = os.path.join(carpeta_salida, nombre_salida)

        # Guardar el archivo combinado en formato .xlsx
        df_combinado.to_excel(ruta_salida, index=False, engine='openpyxl')

        # Ajustar el ancho de la columna `Cuit/Cuil`
        ajustar_ancho_columna(ruta_salida, columna=2, ancho=20)  # Columna 2 es `Cuit/Cuil`

        print(f"Guardado en: {ruta_salida}")
    else:
        print("No se encontraron archivos válidos para procesar.")

# Rutas configurables
carpeta_entrada = r"C:\Users\JUAN49323\OneDrive - Willis Towers Watson\Comisiones mensuales por aseguradora\GALENO\Herramienta para Convertir\Liquidaciones"
carpeta_salida = r"C:\Users\JUAN49323\OneDrive - Willis Towers Watson\Comisiones mensuales por aseguradora\GALENO\2025"

procesar_archivos_excel(carpeta_entrada, carpeta_salida)