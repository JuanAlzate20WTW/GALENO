import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, numbers


def ajustar_ancho_columna(filepath, columna, ancho):
    wb = load_workbook(filepath)
    ws = wb.active
    col_letter = get_column_letter(columna)
    ws.column_dimensions[col_letter].width = ancho

    # Ajustar el formato de la columna como número con decimales
    for row in ws.iter_rows(min_col=columna, max_col=columna, min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.number_format = numbers.FORMAT_NUMBER_00  # Formato con 2 decimales
            cell.alignment = Alignment(horizontal='right')  # Alinear a la derecha

    wb.save(filepath)


def procesar_archivos_excel(carpeta_entrada, carpeta_salida):
    # Crear carpeta de salida si no existe
    os.makedirs(carpeta_salida, exist_ok=True)

    # Obtener mes y año actuales
    fecha_actual = datetime.now()
    mes_actual = fecha_actual.strftime('%B')  # Nombre del mes en español
    anio_actual = fecha_actual.year

    # Calcular el mes anterior
    mes_anterior = (fecha_actual.replace(day=1) - pd.DateOffset(months=1)).strftime('%B')  # Nombre del mes anterior en español

    # Lista para almacenar los datos combinados
    datos_combinados = []

    # Procesar todos los archivos .xls en la carpeta de entrada
    for i, archivo in enumerate(os.listdir(carpeta_entrada)):
        if archivo.lower().endswith('.xls'):
            ruta_entrada = os.path.join(carpeta_entrada, archivo)
            print(f"Procesando: {ruta_entrada}")

            # Leer el archivo .xls
            try:
                # Ajusta el parámetro `header` si los encabezados están en otra fila
                df = pd.read_excel(ruta_entrada, engine='xlrd', header=0)  # Cambia `header=0` si los encabezados están en la primera fila
                print(f"Nombres de columnas en {archivo}: {df.columns.tolist()}")

                # Seleccionar solo las columnas relevantes
                columnas_relevantes = ['Poliza', 'Capitas', 'Masa Salarial']
                df = df[columnas_relevantes]

                # Agregar los datos al conjunto combinado
                datos_combinados.append(df)
            except Exception as e:
                print(f"Error al leer {archivo}: {e}")
                continue

    # Combinar todos los datos en un único DataFrame
    if datos_combinados:
        df_combinado = pd.concat(datos_combinados, ignore_index=True)

        # Eliminar la primera fila si está vacía
        if df_combinado.iloc[0].isnull().all():
            df_combinado = df_combinado.iloc[1:]

        # Generar nombre de salida .xlsx
        nombre_salida = f"GALENO - Premio final - {mes_anterior} {anio_actual}.xlsx"
        ruta_salida = os.path.join(carpeta_salida, nombre_salida)

        # Guardar el archivo combinado en formato .xlsx
        df_combinado.to_excel(ruta_salida, index=False, engine='openpyxl')

        # Cambiar puntos por comas en la columna `Masa Salarial`
        df_combinado['Masa Salarial'] = df_combinado['Masa Salarial'].astype(str).str.replace('.', ',', regex=False)

        # Ajustar el ancho y formato de la columna `Masa Salarial` (columna 3)
        ajustar_ancho_columna(ruta_salida, columna=3, ancho=20)  # Columna 3 es `Masa Salarial`

        print(f"Guardado en: {ruta_salida}")
    else:
        print("No se encontraron archivos válidos para procesar.")


# Rutas configurables
carpeta_entrada = r"C:\Users\JUAN49323\OneDrive - Willis Towers Watson\Comisiones mensuales por aseguradora\GALENO\Herramienta para Convertir\Deudores Por Premio"
carpeta_salida = r"C:\Users\JUAN49323\OneDrive - Willis Towers Watson\Comisiones mensuales por aseguradora\GALENO\2025"

procesar_archivos_excel(carpeta_entrada, carpeta_salida)